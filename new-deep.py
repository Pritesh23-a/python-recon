import numpy as np
import pandas as pd
import pandas as pd
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill, Side, Border

# Load Excel files
bbg_df = pd.read_excel("bbg.xlsx")
murex_df = pd.read_excel("murex.xlsx")
mapping_df = pd.read_excel("MAPPING.xlsx")


# Dictionary for deal mapping: BBG ID → Murex ID
deal_id = dict(zip(mapping_df['BBG_deal_ID'], mapping_df['Murex_Deal_ID']))
# Dictionary for counterparty mapping: murex Name → BBG Name
counterparty_mapping_dictionary = dict(
    zip(mapping_df['COUNTERPARTY_Murex'].str.strip(), mapping_df['COUNTERPARTY_BBG'].str.strip()))

print(counterparty_mapping_dictionary)
# Sorting BBG DataFrame based on mapping BBG ID key
bbg_df_sorted = bbg_df.sort_values(
    by=['Deal ID'], key=lambda x: pd.Categorical(x, categories=deal_id.keys(), ordered=True)
)

# Sorting Murex DataFrame based on mapping Murex ID
murex_df_sorted = murex_df.sort_values(
    by=['Deal ID'], key=lambda x: pd.Categorical(x, categories=list(deal_id.values()), ordered=True)
)

# Adding new column in BBG DataFrame for mapped Murex Deal ID
bbg_df_sorted['murex_Deal_id'] = bbg_df_sorted['Deal ID'].map(deal_id).fillna(np.nan)

print("bbg df sorted--------")
print(bbg_df_sorted)

#TODO
# Drop rows where 'murex_Deal_id' is NaN to avoid merge issues
#bbg_df_sorted = bbg_df_sorted.dropna(subset=['murex_Deal_id'])

# Merging both DataFrames on mapped deal IDs
merge_df = bbg_df_sorted.merge(
    murex_df_sorted, left_on='murex_Deal_id', right_on='Deal ID', indicator=True, suffixes=('_bbg', '_murex')
)

total_col = ['COUNTERPARTY','WE', 'TRANSACTION TYPE', 'PRODUCT','Trade Date','Fixed Rate (%)']

unmatched_bbg_row = bbg_df_sorted[~bbg_df_sorted['murex_Deal_id'].isin(merge_df['murex_Deal_id'])]
unmatched_murex_row=murex_df_sorted[~murex_df_sorted['Deal ID'].isin(merge_df['murex_Deal_id'])]
unmatched_murex_row=unmatched_murex_row.astype(str)
unmatched_bbg_row=unmatched_bbg_row.astype(str)
#unmatched_bbg_murex.loc[:,total_col] =  unmatched_bbg_murex[total_col].astype(str).replace("nan", "#NA")
#unmatched_murex_row.loc[:,total_col] =  unmatched_murex_row[total_col].astype(str).replace("nan", "#NA")
unmatched_bbg_row.loc[:, total_col] = "#NA"
unmatched_murex_row.loc[:, total_col] = "#NA"
print("unmatched bbg df+++++++++++")
print(unmatched_bbg_row)
print("unmatched_murex_row+++++++++++")
print(unmatched_murex_row)
unmapped_columns_to_compare = ['WE', 'TRANSACTION TYPE', 'PRODUCT','Trade Date','Fixed Rate (%)']

for col in unmapped_columns_to_compare:
    bbg_col, murex_col = f'{col}_bbg', f'{col}_murex'

    # Ensure both columns exist in the DataFrame before transformation
    if bbg_col in merge_df.columns and murex_col in merge_df.columns:
        merge_df[murex_col] = merge_df[murex_col].astype(str).str.strip()
        merge_df[bbg_col] = merge_df[bbg_col].astype(str).str.strip()
        merge_df[f'NEW_{col}'] = merge_df.apply(
            lambda row: (
                f"{row[murex_col].strip()}~{row[bbg_col].strip()}"  # If values are different
                if pd.notna(row[bbg_col]) and pd.notna(row[murex_col]) and row[bbg_col] != row[murex_col]
                else  f"{row[murex_col]}-{row[bbg_col]}"   # If values are same

                if row[bbg_col] and row[murex_col] !='nan'
                else '#NA'  # Assign "#NA" if either value is missing
                # if pd.notna(row[bbg_col]) and pd.notna(row[murex_col])
                # else '#NA'  # Assign "#NA" if either value is missing
            ),
            axis=1
        )




print("----Transformed unmapped columns DataFrame-----------")
print(merge_df)

mapped_col_to_compare = ["COUNTERPARTY"]

for col in mapped_col_to_compare:
    bbg_col, murex_col = f'{col}_bbg', f'{col}_murex'

    # Ensure both columns exist in the DataFrame before transformation
    if bbg_col in merge_df.columns and murex_col in merge_df.columns:
        merge_df[murex_col] = merge_df[murex_col].astype(str).str.strip()
        merge_df[bbg_col] = merge_df[bbg_col].astype(str).str.strip()
        merge_df[f'NEW_{col}'] = merge_df.apply(
            lambda row: (
                f"{row[murex_col].strip()} ~ {row[bbg_col].strip()}"  # If values are different
                if pd.notna(row[bbg_col]) and pd.notna(row[murex_col]) and row[bbg_col] != counterparty_mapping_dictionary.get(row[murex_col])
                else  f"{row[murex_col]}-{row[bbg_col]}"   # If values are same
                if row[bbg_col] and row[murex_col] !='nan'
                else '#NA'  # Assign "#NA" if either value is missing
                # if pd.notna(row[bbg_col]) and pd.notna(row[murex_col])
                # else '#NA'  # Assign "#NA" if either value is missing
            ),
            axis=1
        )
print(type(merge_df))
merge_df.columns=merge_df.columns.str.upper()
final_total_col_new = ['DEAL ID_BBG','MUREX_DEAL_ID','NEW_WE', 'NEW_TRANSACTION TYPE', 'NEW_PRODUCT','NEW_TRADE DATE','NEW_FIXED RATE (%)','NEW_COUNTERPARTY','CONTRACT ID']
final_total_col = ['COUNTERPARTY','WE', 'TRANSACTION TYPE', 'PRODUCT','Deal ID','Trade Date','Fixed Rate (%)','DEAL ID_BBG','MUREX_DEAL_ID']

print(merge_df.columns)
merge_df=merge_df[final_total_col_new]


merge_df.columns=merge_df.columns.str.replace('NEW_','')
unmatched_murex_row = unmatched_murex_row.rename(columns={"Deal ID": "MUREX_DEAL_ID"})
unmatched_murex_row.columns=unmatched_murex_row.columns.str.upper()
# merge_df=pd.merge(merge_df,unmatched_murex_row,on='MUREX_DEAL_ID',how='outer')
merge_df = pd.concat([merge_df, unmatched_murex_row], axis=0, ignore_index=True, sort=False)
# merge_df = pd.concat([merge_df, unmatched_murex_row.reindex(columns=merge_df.columns)], ignore_index=True)


new_col_seq = ['DEAL ID_BBG','MUREX_DEAL_ID','COUNTERPARTY','WE', 'TRANSACTION TYPE', 'PRODUCT','TRADE DATE','FIXED RATE (%)','CONTRACT ID']
# new_merge_df=merge_df[[new_col_seq]]
#working TODO
#merge_df = pd.concat([merge_df, unmatched_murex_row]).drop_duplicates().reset_index(drop=True)

# new_merge_df.to_excel("output-merge_new.xlsx", index=False)
merge_df.to_excel("output-merge.xlsx", index=False)



# Sample DataFrame

# Write DataFrame to Excel
file_path = "recon_output.xlsx"
merge_df.to_excel(file_path, sheet_name="Recon", index=False)

# Open the workbook with openpyxl
wb = load_workbook(file_path)
ws = wb["Recon"]

# Define fill colors
green_fill = PatternFill(start_color="00FF00", end_color="DAFFD5", fill_type="lightTrellis")  # Green for "_"
red_fill = PatternFill(start_color="FF0000", end_color="FA6B84", fill_type="solid")  # Red for "~"

# Apply color formatting based on **cell values**
# Define border style
thin_border = Border(left=Side(style="thin"),
                     right=Side(style="thin"),
                     top=Side(style="thin"),
                     bottom=Side(style="thin"))

# Apply color formatting and border to **data cells**
for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=2):
    for cell in row:
        cell.border = thin_border  # Apply border to every cell
        if "~" in str(cell.value) or " ~ " in str(cell.value):  # Check if value contains "_"
            cell.fill = red_fill
        elif "-" in str(cell.value) :  # Check if value contains "~"
            cell.fill = green_fill

# Save the updated workbook
wb.save(file_path)
print(f"Excel file saved as {file_path}")